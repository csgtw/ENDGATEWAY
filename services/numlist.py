"""
services/numlist.py
Gestion de la numlist : import CSV/XLSX, stockage Redis, draft message.
Multi-lists : chaque fichier importé crée une liste nommée indépendante (nl:list:{id}).
"""
import csv
import json
import time
import re
import uuid
from io import BytesIO, StringIO

from openpyxl import load_workbook
from services.redis_store import redis_conn

NL_QUEUE_KEY  = "nl:queue"   # legacy — conservé pour compat
NL_META_KEY   = "nl:meta"
NL_DRAFT_KEY  = "nl:draft"
NL_LISTS_KEY  = "nl:lists"   # Hash : list_id → JSON metadata
NL_SELECTED_KEY = "nl:selected"  # JSON list d'IDs sélectionnés, absent = tous

BATCH_SIZE = 1500  # Contacts par pipeline rpush


def _now() -> int:
    return int(time.time())


def _clean_header(h) -> str:
    h = str(h or "").strip()
    h = re.sub(r"\s+", "_", h)
    return h


def _is_probably_phone_header(h: str) -> bool:
    return h.strip().lower() in (
        "number", "phone", "tel", "telephone", "mobile", "msisdn", "num", "numero"
    )


def _normalize_number(s: str) -> str:
    s = str(s or "").strip()
    if not s:
        return ""
    s = re.sub(r"[^\d\+]", "", s)
    if s.count("+") > 1:
        s = "+" + re.sub(r"[^\d]", "", s)
    if s.startswith("00"):
        s = "+" + s[2:]
    return s


def _detect_number_col(headers: list) -> str:
    for h in headers:
        if _is_probably_phone_header(h):
            return h
    return headers[0] if headers else "number"


def _iter_xlsx(file_bytes: bytes):
    """Générateur streaming : yield (headers, row_dict) ligne par ligne — jamais toute la liste en mémoire."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = None
    try:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [_clean_header(x) for x in row]
                headers = [h or f"col_{idx+1}" for idx, h in enumerate(headers)]
                continue
            if headers is None:
                continue
            obj = {headers[idx]: (row[idx] if idx < len(row) else None) for idx in range(len(headers))}
            yield headers, obj
    finally:
        wb.close()


def _iter_txt(file_bytes: bytes):
    """Générateur pour fichiers .txt : un numéro par ligne, pas de header."""
    try:
        text = file_bytes.decode("utf-8")
    except Exception:
        text = file_bytes.decode("latin-1", errors="ignore")
    headers = ["number"]
    for line in text.splitlines():
        line = line.strip()
        if line:
            yield headers, {"number": line}


def _iter_csv(file_bytes: bytes):
    """Générateur streaming : yield (headers, row_dict) ligne par ligne — jamais toute la liste en mémoire."""
    try:
        text = file_bytes.decode("utf-8")
    except Exception:
        text = file_bytes.decode("latin-1", errors="ignore")

    reader = csv.reader(StringIO(text))
    headers = None
    for i, row in enumerate(reader):
        if i == 0:
            headers = [_clean_header(x) for x in row]
            headers = [h or f"col_{idx+1}" for idx, h in enumerate(headers)]
            continue
        if headers is None:
            continue
        obj = {headers[idx]: (row[idx] if idx < len(row) else None) for idx in range(len(headers))}
        yield headers, obj


def get_selected_lists() -> list | None:
    """Retourne la liste des IDs de listes sélectionnées. None = toutes sélectionnées."""
    try:
        v = redis_conn.get(NL_SELECTED_KEY)
        if not v:
            return None
        data = json.loads(v.decode("utf-8") if isinstance(v, bytes) else v)
        return data if isinstance(data, list) else None
    except Exception:
        return None


def set_selected_lists(list_ids: list | None):
    """Définit les listes sélectionnées. None = toutes sélectionnées."""
    if list_ids is None:
        redis_conn.delete(NL_SELECTED_KEY)
    else:
        redis_conn.set(NL_SELECTED_KEY, json.dumps([str(x) for x in list_ids]))


def _meta_imported_total() -> int:
    try:
        return int(redis_conn.hget(NL_META_KEY, "imported_total") or 0)
    except Exception:
        return 0


def _import_single_file(f) -> dict:
    """
    Importe un seul fichier dans une liste nommée dédiée (nl:list:{id}).
    Retourne {"list_id", "added", "variables", "number_col"}.
    """
    list_id  = str(uuid.uuid4())[:8]
    list_key = f"nl:list:{list_id}"
    filename = f.filename or "import"

    raw_name = (f.filename or "").lower()
    content  = f.read()
    if raw_name.endswith(".xlsx"):
        row_iter = _iter_xlsx(content)
    elif raw_name.endswith(".csv"):
        row_iter = _iter_csv(content)
    elif raw_name.endswith(".txt"):
        row_iter = _iter_txt(content)
    else:
        raise ValueError(f"Format non supporté pour '{f.filename}' (xlsx/csv/txt uniquement)")

    file_headers    = None
    file_number_col = None
    variables       = set()
    batch           = []
    total_added     = 0

    def flush():
        nonlocal batch
        if not batch:
            return
        pipe = redis_conn.pipeline()
        pipe.rpush(list_key, *batch)
        pipe.execute()
        batch = []

    for headers, row in row_iter:
        if file_headers is None:
            file_headers    = headers
            file_number_col = _detect_number_col(headers)
            for h in headers:
                if h != file_number_col:
                    variables.add(h)

        raw_num = row.get(file_number_col)
        number  = _normalize_number(raw_num)
        if not number:
            continue

        contact = {"number": number}
        for k, v in row.items():
            if k == file_number_col:
                continue
            kk = _clean_header(k)
            if not kk:
                continue
            contact[kk] = "" if v is None else str(v).strip()

        batch.append(json.dumps(contact, ensure_ascii=False))
        total_added += 1
        if len(batch) >= BATCH_SIZE:
            flush()

    flush()

    # Enregistrer dans le registre
    meta = {
        "name":       filename,
        "created_ts": _now(),
        "number_col": file_number_col or "number",
        "variables":  sorted(list(variables)),
    }
    redis_conn.hset(NL_LISTS_KEY, list_id, json.dumps(meta, ensure_ascii=False))

    return {
        "list_id":    list_id,
        "added":      total_added,
        "variables":  sorted(list(variables)),
        "number_col": file_number_col or "number",
    }


def import_files(files) -> dict:
    """
    files : list de Werkzeug FileStorage.
    Chaque fichier crée une liste nommée indépendante.
    Retourne {"added", "imported_total", "variables", "number_col"}.
    """
    total_added = 0
    all_variables: set = set()
    chosen_number_col = "number"

    for f in files:
        res = _import_single_file(f)
        total_added      += res["added"]
        all_variables    |= set(res["variables"])
        chosen_number_col = res["number_col"]

    new_total = int(_meta_imported_total()) + total_added
    meta = {
        "number_col":     chosen_number_col,
        "variables":      json.dumps(sorted(list(all_variables)), ensure_ascii=False),
        "imported_total": str(new_total),
        "updated_ts":     str(_now()),
    }
    redis_conn.hset(NL_META_KEY, mapping=meta)

    return {
        "added":          total_added,
        "imported_total": new_total,
        "variables":      sorted(list(all_variables)),
        "number_col":     chosen_number_col,
    }


def get_named_lists() -> list:
    """Retourne toutes les listes nommées avec leur count Redis actuel."""
    try:
        raw = redis_conn.hgetall(NL_LISTS_KEY)
        if not raw:
            return []
        items = []
        pipe  = redis_conn.pipeline()
        for k, v in raw.items():
            lid  = k.decode("utf-8") if isinstance(k, bytes) else k
            meta = json.loads(v.decode("utf-8") if isinstance(v, bytes) else v)
            items.append((lid, meta))
            pipe.llen(f"nl:list:{lid}")
        counts = pipe.execute()
        result = []
        for i, (lid, meta) in enumerate(items):
            result.append({
                "id":         lid,
                "name":       meta.get("name", lid),
                "created_ts": meta.get("created_ts", 0),
                "count":      counts[i] if i < len(counts) else 0,
            })
        result.sort(key=lambda x: x["created_ts"])
        return result
    except Exception:
        return []


def get_list_contacts(list_id: str, offset: int = 0, limit: int = 200) -> dict:
    """
    Lit les contacts d'une liste spécifique avec pagination.
    Retourne {total, contacts, columns}.
    """
    list_id = str(list_id).strip()
    list_key = f"nl:list:{list_id}"
    try:
        total = int(redis_conn.llen(list_key) or 0)
        if total == 0:
            return {"total": 0, "contacts": [], "columns": []}
        # lrange est de gauche (index 0) à droite (index n-1)
        # les contacts sont dépilés par la droite (rpop) donc la "tête" est à droite
        start = max(0, total - offset - limit)
        end   = max(0, total - offset - 1)
        if start > end:
            return {"total": total, "contacts": [], "columns": []}
        rows = redis_conn.lrange(list_key, start, end) or []
        contacts = []
        columns_seen = set()
        columns = []
        for raw in reversed(rows):
            try:
                c = json.loads(raw.decode("utf-8") if isinstance(raw, bytes) else raw)
                for k in c:
                    if k not in columns_seen:
                        columns_seen.add(k)
                        columns.append(k)
                contacts.append(c)
            except Exception:
                continue
        return {"total": total, "contacts": contacts, "columns": columns}
    except Exception:
        return {"total": 0, "contacts": [], "columns": []}


def delete_contact_from_list(list_id: str, number: str) -> bool:
    """Supprime la première occurrence d'un contact (par numéro) dans une liste."""
    list_id = str(list_id).strip()
    number  = str(number).strip()
    if not list_id or not number:
        return False
    list_key = f"nl:list:{list_id}"
    try:
        # Trouver et supprimer la première occurrence
        total = int(redis_conn.llen(list_key) or 0)
        for raw in (redis_conn.lrange(list_key, 0, total - 1) or []):
            try:
                c = json.loads(raw.decode("utf-8") if isinstance(raw, bytes) else raw)
                if str(c.get("number", "")).strip() == number:
                    redis_conn.lrem(list_key, 1, raw)
                    return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def delete_named_list(list_id: str):
    """Supprime une liste nommée et tous ses contacts."""
    list_id = str(list_id).strip()
    if not list_id:
        return
    redis_conn.delete(f"nl:list:{list_id}")
    redis_conn.hdel(NL_LISTS_KEY, list_id)


def pop_contact_from_lists() -> tuple:
    """
    Pop le prochain contact depuis la première liste non-vide.
    Retourne (bytes|None, list_key|None) pour permettre le re-push vers la bonne liste.
    Fallback sur nl:queue (legacy).
    Respecte la sélection de listes (nl:selected).
    """
    try:
        selected = get_selected_lists()
        raw_ids = redis_conn.hkeys(NL_LISTS_KEY)
        for raw_id in (raw_ids or []):
            lid = raw_id.decode("utf-8") if isinstance(raw_id, bytes) else raw_id
            if selected is not None and lid not in selected:
                continue
            list_key = f"nl:list:{lid}"
            val = redis_conn.rpop(list_key)
            if val:
                return val, list_key
    except Exception:
        pass
    # Fallback legacy
    try:
        val = redis_conn.rpop(NL_QUEUE_KEY)
        return val, NL_QUEUE_KEY if val else None
    except Exception:
        return None, None


def peek_contacts_from_lists(count: int) -> list:
    """
    Retourne les `count` prochains contacts sans les dépiler.
    Lit depuis toutes les listes nommées puis fallback legacy.
    """
    result = []
    try:
        raw_ids = redis_conn.hkeys(NL_LISTS_KEY)
        for raw_id in (raw_ids or []):
            if len(result) >= count:
                break
            lid  = raw_id.decode("utf-8") if isinstance(raw_id, bytes) else raw_id
            need = count - len(result)
            n    = redis_conn.llen(f"nl:list:{lid}")
            if not n:
                continue
            take  = min(need, n)
            start = max(0, n - take)
            rows  = redis_conn.lrange(f"nl:list:{lid}", start, n - 1) or []
            for raw in reversed(rows):
                try:
                    result.append(json.loads(raw.decode("utf-8")))
                except Exception:
                    continue
    except Exception:
        pass

    # Fallback legacy nl:queue si pas assez
    if len(result) < count:
        need  = count - len(result)
        try:
            n = redis_conn.llen(NL_QUEUE_KEY)
            if n:
                start = max(0, n - need)
                rows  = redis_conn.lrange(NL_QUEUE_KEY, start, n - 1) or []
                for raw in reversed(rows):
                    try:
                        result.append(json.loads(raw.decode("utf-8")))
                    except Exception:
                        continue
        except Exception:
            pass

    return result[:count]


def load_nl_meta() -> dict | None:
    m = redis_conn.hgetall(NL_META_KEY) or {}
    if not m:
        return None
    out = {}
    for k, v in m.items():
        kk = k.decode("utf-8", errors="ignore")
        vv = v.decode("utf-8", errors="ignore")
        out[kk] = vv
    try:
        out["variables"] = json.loads(out.get("variables") or "[]")
    except Exception:
        out["variables"] = []
    try:
        out["imported_total"] = int(out.get("imported_total") or 0)
    except Exception:
        out["imported_total"] = 0
    return out


def nl_remaining_count() -> int:
    try:
        selected = get_selected_lists()
        total = 0
        raw_ids = redis_conn.hkeys(NL_LISTS_KEY)
        if raw_ids:
            pipe = redis_conn.pipeline()
            active_ids = []
            for raw_id in raw_ids:
                lid = raw_id.decode("utf-8") if isinstance(raw_id, bytes) else raw_id
                if selected is not None and lid not in selected:
                    continue
                active_ids.append(lid)
                pipe.llen(f"nl:list:{lid}")
            if active_ids:
                counts = pipe.execute()
                total = sum(c or 0 for c in counts)
        legacy = int(redis_conn.llen(NL_QUEUE_KEY) or 0)
        return total + legacy
    except Exception:
        return 0


def clear_numlist():
    """Supprime toutes les listes nommées + legacy queue + meta."""
    try:
        raw_ids = redis_conn.hkeys(NL_LISTS_KEY) or []
        if raw_ids:
            pipe = redis_conn.pipeline()
            for raw_id in raw_ids:
                lid = raw_id.decode("utf-8") if isinstance(raw_id, bytes) else raw_id
                pipe.delete(f"nl:list:{lid}")
            pipe.execute()
        redis_conn.delete(NL_LISTS_KEY)
    except Exception:
        pass
    redis_conn.delete(NL_QUEUE_KEY)
    redis_conn.delete(NL_META_KEY)


def save_message_draft(message: str, msg_type: str):
    msg_type = (msg_type or "sms").strip().lower()
    if msg_type not in ("sms", "mms"):
        msg_type = "sms"
    redis_conn.hset(NL_DRAFT_KEY, mapping={
        "message":    str(message or ""),
        "type":       msg_type,
        "updated_ts": str(_now()),
    })


def load_message_draft() -> tuple[str, str]:
    try:
        m = redis_conn.hgetall(NL_DRAFT_KEY) or {}
        msg = (m.get(b"message") or b"").decode("utf-8", errors="ignore")
        typ = (m.get(b"type") or b"sms").decode("utf-8", errors="ignore")
        typ = typ if typ in ("sms", "mms") else "sms"
        return msg, typ
    except Exception:
        return "", "sms"
